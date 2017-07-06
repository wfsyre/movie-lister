import imdb
import os
import sys
from openpyxl import Workbook, load_workbook
import time

file_extensions = [".avi", ".mkv", ".mp4", ".m4v", ".mov"]
found_movies = []
movie_info = {}
ErrLog = []
def compareStrings(s1, s2, count):
    for c1, c2 in zip(s1, s2):
        if c1 != c2:
            count -= 1
            if count <= 0:
                return False
    return True
start_time = time.time()
imdb_access = imdb.IMDb()
try:
    wb = load_workbook(filename="movie info.xlsx")
    for row in wb['Movies']:
        movie = []
        for cell in row:
            movie.append(cell.value)
        if len(movie) == 5:
            movie_info[movie[0]] = (movie[1], movie[2], movie[3], movie[4])
except Exception:
    print "Could not load movie info.xlsx"
c = 1
collected_movies = movie_info.keys()
for filename in os.listdir("."):
    print "\r",
    done_string = "[                    ]" #20 spaces
    for i in range(1, 21):
        if (float(c) / len(os.listdir(".")) * 100) >= i * 5:
            done_string = done_string[0:i] + "#" + done_string[i + 1:]
    print "%s %s %.2f%% %-100s" % ("Processing Movies", done_string, float(c) / len(os.listdir(".")) * 100, filename[0:-10]),
    c += 1
    if c % 700 == 0:
        imdb_access = imdb.IMDb()
    if len(filename) >= 5 and filename[-4:] in file_extensions:
        if filename[0:-11] not in collected_movies:
            try:
                movie_possibilities = imdb_access.search_movie(filename[0:-4])
            except Exception as a:
                ErrMessage = a.message
                ErrLog.append((ErrMessage, filename, a))
                imdb_access = imdb.IMDb()
                movie_possibilities = imdb_access.search_movie(filename[0:-4])
            else:
                imdb_access = imdb.IMDb()
                movie_possibilities = imdb_access.search_movie(filename[0:-4])
            if len(movie_possibilities) > 0:
                try:
                    movie = imdb_access.get_movie(movie_possibilities[0].movieID)
                except Exception as a:
                    ErrMessage = a.message
                    ErrLog.append((ErrMessage, filename, a))
                    imdb_access = imdb.IMDb()
                    movie = imdb_access.get_movie(movie_possibilities[0].movieID)
                if "title" in movie.keys() and movie["title"] not in movie_info.keys():
                    if "genres" in movie.keys():
                        genre_list = movie["genres"]
                        new_genre_list = []
                        for genre in genre_list:
                            new_genre_list.append(genre.encode("ascii"))
                        genre_string = ', '.join(new_genre_list)
                    else:
                        genre_string = "None Found"
                    if "production companies" in movie.keys():
                        pro = movie["production companies"][0]["name"]
                    else:
                        pro = "None Found"
                    movie_info[movie["title"]] = (movie["year"] if "year" in movie.keys() else "None Found", genre_string, movie["rating"] if "rating" in movie.keys() else "None Found", pro)
print "\r",
print "Processing Movies [####################]"
sys.stdout.flush()
sorted_movies = movie_info.keys()
sorted_movies.sort()
new_movie_list = Workbook()
current_sheet = new_movie_list.active
current_sheet.title = "Movies"
movie_number = 0
current_sheet["A1"] = "Title"
current_sheet["B1"] = "Year"
current_sheet["C1"] = "Genre(s)"
current_sheet["D1"] = "Rating (IMDb)"
current_sheet["E1"] = "Production Company"
for row in range(2, 1 + len(sorted_movies)):
    count = -1
    for string in ["A" + str(row), "B" + str(row), "C" + str(row), "D" + str(row), "E" + str(row)]:
        if count == -1:
            current_sheet[string] = sorted_movies[movie_number]
            count += 1
        else:
            current_sheet[string] = movie_info[sorted_movies[movie_number]][count]
            count += 1
    movie_number += 1
new_movie_list.save("movie info.xlsx")
print "Excel Sheet Generated"
print "%d movies done in %.1f seconds" % (movie_number + 1, time.time() - start_time)
print "Errors:"
print ErrLog
